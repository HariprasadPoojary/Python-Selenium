from Basics.test_framework.utilities.logger import InfoLogger
import openpyxl

class Utilities:

    log = InfoLogger('debug').info_logger()

    def __init__(self, excel_file=None, excel_sheet=None):
        if excel_file is not None and excel_sheet is not None:
            self.excel_file = excel_file
            self.excel_sheet = excel_sheet
            self.work_book = openpyxl.load_workbook(self.excel_file)
            self.work_sheet = self.work_book[self.excel_sheet]
            self.log.info(f'Utilities class instantiated with excel file where filename - {self.excel_file}'
                          f'Sheet name - {self.excel_sheet}')
        else:
            self.log.info(f'Utilities class instantiated without any parameters')

    def excel_rowcount(self):
        rowcount = self.work_sheet.max_row
        self.log.info(f'Row count for excel file {self.excel_file} is {rowcount}')
        return rowcount

    def excel_columncount(self):
        colcount = self.work_sheet.max_column
        self.log.info(f'Column count for excel file {self.excel_file} is {colcount}')
        return colcount

    def excel_readcell(self, row_num, col_num):
        cell_value = self.work_sheet.cell(row=row_num, column=col_num)
        self.log.info(f'Extracted value {cell_value} from excel file - {self.excel_file}, sheet - {self.excel_sheet}')
        return cell_value

    def excel_writecell(self, row_num, col_num, data):
        self.work_sheet.cell(row=row_num, column=col_num).value = data
        self.log.info(f'{data} written to excel file - {self.excel_file}, sheet - {self.excel_sheet}, row - {row_num}'
                      f'column - {col_num}')
        self.work_book.save(self.excel_file)
        self.log.info(f'Excel file - {self.excel_file} has been saved')

    def excel_getlist_cellvalues(self, min_row=1, max_row=2, min_col=1, max_col=2):
        try:
            if (max_row > min_row) and (max_col > min_col):
                data_rows = (max_row - min_row) + 1
            else:
                data_rows = 1

            data_list = [[] for i in range(data_rows)]
            list_counter = 0

            for row in self.work_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    data_list[list_counter].append(cell.value)
                list_counter += 1

            return data_list
        except Exception as err:
            self.log.error(f'Exception - {err} occurred while getting a list of cell values from '
                           f'Excel - {self.excel_file} Excel sheet - {self.excel_sheet}')
